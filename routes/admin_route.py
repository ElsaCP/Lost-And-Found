from flask import (Blueprint, render_template, request, jsonify, redirect, url_for, session, flash, current_app, send_file)
import mysql.connector
import os
import time
import re
import io
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image as RLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from flask_mail import Message
from extensions import mail
from PIL import Image


admin_bp = Blueprint(
    'admin_bp',
    __name__,
    static_folder='../static_admin',
    static_url_path='/static_admin',
    template_folder='../templates/admin'
)

def get_db_connection():
    return mysql.connector.connect(
        host='localhost',
        user='root',         
        password='',         
        database='lostfound'  
    )

@admin_bp.route('/login', methods=['GET', 'POST'])
def login_admin():
    if request.method == 'GET':
        return render_template('index.html')

    data = request.get_json()
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()

    if not email or not password:
        return jsonify({
            'success': False,
            'message': 'Email dan password wajib diisi'
        }), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id, full_name, email, phone, role, password
        FROM admin
        WHERE email=%s
    """, (email,))

    admin = cursor.fetchone()
    cursor.close()
    conn.close()

    if admin and check_password_hash(admin['password'], password):
        session['admin_logged_in'] = True
        session['admin_email'] = admin['email']
        session['admin_id'] = admin['id']
        session['role'] = admin['role']

        return jsonify({
            'success': True,
            'role': admin['role']
        })

    return jsonify({
        'success': False,
        'message': 'Email atau password salah!'
    }), 401

@admin_bp.route('/beranda', endpoint='beranda_admin')
def beranda_admin():
    auto_arsip_laporan()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        (
            SELECT 
                id,
                kode_kehilangan AS kode,
                nama_barang,
                nama_pelapor,
                no_telp,
                email,
                lokasi,
                tanggal_kehilangan AS tanggal,
                status,
                update_terakhir,
                'kehilangan' AS jenis_laporan
            FROM kehilangan
            WHERE is_arsip = 0
        )

        UNION ALL

        (
            SELECT 
                id,
                kode_barang AS kode,
                nama_barang,
                nama_pelapor,
                no_telp,
                email,
                lokasi,
                tanggal_lapor AS tanggal,
                status,
                update_terakhir,
                'penemuan' AS jenis_laporan
            FROM penemuan
            WHERE is_arsip = 0
        )

        UNION ALL

        (
            SELECT 
                id,
                kode_laporan AS kode,
                nama_barang,
                nama_pelapor,
                no_telp,
                email,
                NULL AS lokasi,
                tanggal_lapor AS tanggal,
                status,
                update_terakhir,
                'klaim' AS jenis_laporan
            FROM klaim_barang
            WHERE is_arsip = 0
        )

        ORDER BY update_terakhir DESC
    """

    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template("admin/beranda.html", data=data)

def format_bulan_indonesia(bulan_str):
    bulan_map = {
        "01": "Januari", "02": "Februari", "03": "Maret",
        "04": "April", "05": "Mei", "06": "Juni",
        "07": "Juli", "08": "Agustus", "09": "September",
        "10": "Oktober", "11": "November", "12": "Desember"
    }

    tahun, bulan = bulan_str.split("-")
    return f"{bulan_map.get(bulan, bulan)} {tahun}"

def compress_image(file, output_path, max_size=(1280, 1280), quality=70):
    image = Image.open(file)

    if image.mode in ("RGBA", "P"):
        image = image.convert("RGB")

    image.thumbnail(max_size)

    image.save(output_path, format="JPEG", quality=quality, optimize=True)

@admin_bp.route('/export/pdf')
def export_pdf():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')  
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT kode_kehilangan AS kode, nama_barang, 'kehilangan' AS jenis,
               status, deskripsi AS deskripsi_text, lokasi, DATE_FORMAT(tanggal_kehilangan, '%d-%m-%Y') AS tanggal,
               foto AS foto_barang
        FROM kehilangan
        WHERE DATE_FORMAT(tanggal_kehilangan, '%Y-%m') = %s
        UNION ALL
        SELECT kode_barang AS kode, nama_barang, 'penemuan' AS jenis,
               status, deskripsi AS deskripsi_text, lokasi, DATE_FORMAT(tanggal_lapor, '%d-%m-%Y') AS tanggal,
               gambar_barang AS foto_barang
        FROM penemuan
        WHERE DATE_FORMAT(tanggal_lapor, '%Y-%m') = %s
        UNION ALL
        SELECT kode_laporan AS kode, nama_barang, 'klaim' AS jenis,
               status, deskripsi_khusus AS deskripsi_text, '-' AS lokasi, DATE_FORMAT(tanggal_lapor, '%d-%m-%Y') AS tanggal,
               foto_barang
        FROM klaim_barang
        WHERE DATE_FORMAT(tanggal_lapor, '%Y-%m') = %s
        ORDER BY tanggal ASC
    """
    cursor.execute(query, (bulan, bulan, bulan))
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=40
    )
    styles = getSampleStyleSheet()
    normal = styles['Normal']

    elements = [Paragraph(f"<b>Rekapan Laporan Bulan {bulan_label}</b>", styles['Title'])]

    table_data = [["Kode", "Nama Barang", "Jenis", "Status", "Deskripsi", "Lokasi", "Tanggal", "Foto"]]

    for row in data:
        foto_cell = Paragraph("-", normal)
        if row.get("foto_barang"):
            foto_path = os.path.join(
                current_app.root_path, "static", "uploads", row["foto_barang"]
            )
            if os.path.exists(foto_path):
                foto_cell = RLImage(foto_path, width=50, height=50)

        deskripsi = row.get("deskripsi_text") or "-"

        table_data.append([
            Paragraph(str(row.get('kode', '-')), normal),
            Paragraph(str(row.get('nama_barang', '-')), normal),
            Paragraph(str(row.get('jenis', '-')), normal),
            Paragraph(str(row.get('status', '-')), normal),
            Paragraph(deskripsi, normal),
            Paragraph(str(row.get('lokasi', '-')), normal),
            Paragraph(str(row.get('tanggal', '-')), normal),
            foto_cell
        ])

    col_widths = [50, 80, 60, 60, 120, 80, 60, 50]

    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (4,1), (4,-1), 'LEFT'),  
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"rekap_laporan_{bulan_label.replace(' ', '_').lower()}.pdf",
        mimetype="application/pdf"
    )

@admin_bp.route('/export/excel')
def export_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT kode_kehilangan AS kode, nama_barang, 'kehilangan' AS jenis,
               status, deskripsi AS deskripsi_text, lokasi, tanggal_kehilangan AS tanggal
        FROM kehilangan
        WHERE DATE_FORMAT(tanggal_kehilangan, '%Y-%m') = %s
        UNION ALL
        SELECT kode_barang AS kode, nama_barang, 'penemuan' AS jenis,
               status, deskripsi AS deskripsi_text, lokasi, tanggal_lapor AS tanggal
        FROM penemuan
        WHERE DATE_FORMAT(tanggal_lapor, '%Y-%m') = %s
        UNION ALL
        SELECT kode_laporan AS kode, nama_barang, 'klaim' AS jenis,
               status, deskripsi_khusus AS deskripsi_text, '-' AS lokasi, tanggal_lapor AS tanggal
        FROM klaim_barang
        WHERE DATE_FORMAT(tanggal_lapor, '%Y-%m') = %s
        ORDER BY tanggal ASC
    """
    cursor.execute(query, (bulan, bulan, bulan))
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Laporan"

    ws.append(["Kode", "Nama Barang", "Jenis", "Status", "Deskripsi", "Lokasi", "Tanggal"])

    for d in data:
        ws.append([
            d['kode'],
            d['nama_barang'],
            d['jenis'],
            d['status'],
            d.get('deskripsi_text', '-'),
            d.get('lokasi', '-'),
            d['tanggal']
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"rekap_laporan_{bulan_label.replace(' ', '_').lower()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@admin_bp.route('/kehilangan/export/pdf')
def export_kehilangan_pdf():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            kode_kehilangan,
            nama_pelapor,
            nama_barang,
            lokasi,
            status,
            deskripsi,
            DATE_FORMAT(tanggal_submit, '%d-%m-%Y') AS tanggal,
            foto
        FROM kehilangan
        WHERE DATE_FORMAT(tanggal_submit, '%Y-%m') = %s
        ORDER BY tanggal_submit ASC
    """, (bulan,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=30,
        rightMargin=30,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]

    elements = [
        Paragraph(f"<b>Laporan Kehilangan Bulan {bulan_label}</b>", styles["Title"]),
        Paragraph("<br/>", normal)
    ]

    table_data = [[
        Paragraph("Kode", normal),
        Paragraph("Pelapor", normal),
        Paragraph("Barang", normal),
        Paragraph("Lokasi", normal),
        Paragraph("Status", normal),
        Paragraph("Deskripsi", normal),
        Paragraph("Tanggal", normal),
        Paragraph("Foto", normal)
    ]]

    for row in data:
        foto_cell = Paragraph("-", normal)
        if row["foto"]:
            foto_path = os.path.join(
                current_app.root_path,
                "static",
                "uploads",
                row["foto"]
            )
            if os.path.exists(foto_path):
                foto_cell = RLImage(foto_path, width=50, height=50)

        table_data.append([
            Paragraph(row["kode_kehilangan"], normal),
            Paragraph(row["nama_pelapor"], normal),
            Paragraph(row["nama_barang"], normal),
            Paragraph(row["lokasi"], normal),
            Paragraph(row["status"], normal),
            Paragraph(row.get('deskripsi', '-'), normal),  
            Paragraph(row["tanggal"], normal),
            foto_cell
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[55, 70, 70, 90, 60, 100, 60, 50] 
    )

    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),  
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (5,1), (5,-1), 'LEFT'),  
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"laporan_kehilangan_{bulan_label.replace(' ', '_').lower()}.pdf",
        mimetype="application/pdf"
    )
    
@admin_bp.route('/kehilangan/export/excel')
def export_kehilangan_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            kode_kehilangan AS kode,
            nama_pelapor,
            nama_barang,
            lokasi,
            status,
            deskripsi,
            DATE_FORMAT(tanggal_submit, '%d-%m-%Y') AS tanggal
        FROM kehilangan
        WHERE DATE_FORMAT(tanggal_submit, '%Y-%m') = %s
        ORDER BY tanggal_submit ASC
    """, (bulan,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Kehilangan"

    ws.append(["Kode", "Pelapor", "Nama Barang", "Lokasi", "Status", "Deskripsi", "Tanggal"])

    for d in data:
        ws.append([
            d['kode'],
            d['nama_pelapor'],
            d['nama_barang'],
            d.get('lokasi', '-'),
            d['status'],
            d.get('deskripsi', '-'),
            d['tanggal']
        ])

    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            if cell.value:
                cell.number_format = 'DD-MM-YYYY'

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"laporan_kehilangan_{bulan_label.replace(' ', '_').lower()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@admin_bp.route('/penemuan/export/pdf')
def export_penemuan_pdf():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            kode_barang,
            nama_barang,
            lokasi,
            status,
            deskripsi,
            DATE_FORMAT(tanggal_lapor, '%d-%m-%Y') AS tanggal,
            gambar_barang
        FROM penemuan
        WHERE DATE_FORMAT(tanggal_lapor, '%Y-%m') = %s
        ORDER BY tanggal_lapor ASC
    """, (bulan,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=30,
        rightMargin=30,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]

    elements = [
        Paragraph(f"<b>Laporan Penemuan Bulan {bulan_label}</b>", styles["Title"]),
        Paragraph("<br/>", normal)
    ]

    table_data = [[
        "Kode",
        "Barang",
        "Lokasi",
        "Status",
        "Deskripsi",
        "Tanggal",
        "Foto"
    ]]

    for row in data:
        foto_cell = Paragraph("-", normal)
        if row["gambar_barang"]:
            foto_path = os.path.join(
                current_app.root_path,
                "static",
                "uploads",
                row["gambar_barang"]
            )
            if os.path.exists(foto_path):
                foto_cell = RLImage(foto_path, width=50, height=50)

        table_data.append([
            Paragraph(row["kode_barang"], normal),
            Paragraph(row["nama_barang"], normal),
            Paragraph(row["lokasi"], normal),
            Paragraph(row["status"], normal),
            Paragraph(row.get("deskripsi", "-"), normal), 
            Paragraph(row["tanggal"], normal),
            foto_cell
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[60, 100, 100, 70, 120, 60, 50]  
    )

    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),  
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (4,1), (4,-1), 'LEFT'),   
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"laporan_penemuan_{bulan_label.replace(' ', '_').lower()}.pdf",
        mimetype="application/pdf"
    )

@admin_bp.route('/penemuan/export/excel')
def export_penemuan_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            kode_barang AS kode,
            nama_barang,
            lokasi,
            status,
            deskripsi,
            tanggal_lapor AS tanggal
        FROM penemuan
        WHERE DATE_FORMAT(tanggal_lapor, '%Y-%m') = %s
        ORDER BY tanggal_lapor ASC
    """, (bulan,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penemuan"

    ws.append(["Kode", "Nama Barang", "Lokasi", "Status", "Deskripsi", "Tanggal"])

    for d in data:
        ws.append([
            d['kode'],
            d['nama_barang'],
            d.get('lokasi', '-'),
            d['status'],
            d.get('deskripsi', '-'),
            d['tanggal']
        ])

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            if cell.value:
                cell.number_format = 'DD-MM-YYYY'

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"laporan_penemuan_{bulan_label.replace(' ', '_').lower()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@admin_bp.route('/klaim/export/pdf')
def export_klaim_pdf():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')  
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            kode_laporan AS kode_klaim,
            kode_laporan_kehilangan,
            nama_barang,
            nama_pelapor,
            status,
            deskripsi_khusus,
            DATE_FORMAT(STR_TO_DATE(tanggal_lapor, '%Y-%m-%d'), '%d-%m-%Y') AS tanggal,
            foto_barang
        FROM klaim_barang
        WHERE DATE_FORMAT(STR_TO_DATE(tanggal_lapor, '%Y-%m-%d'), '%Y-%m') = %s
        ORDER BY STR_TO_DATE(tanggal_lapor, '%Y-%m-%d') ASC
    """, (bulan,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=30,
        rightMargin=30,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]

    elements = [
        Paragraph(f"<b>Laporan Klaim Barang Bulan {bulan_label}</b>", styles['Title']),
        Paragraph("<br/>", normal)
    ]

    table_data = [[
        Paragraph("Kode<br/>Klaim", normal),
        Paragraph("Kode<br/>Kehilangan", normal),
        Paragraph("Nama Barang", normal),
        Paragraph("Nama Pengklaim", normal),
        Paragraph("Status", normal),
        Paragraph("Deskripsi", normal),
        Paragraph("Tanggal", normal),
        Paragraph("Foto", normal)
    ]]

    for row in data:
        # Foto
        foto_cell = Paragraph("-", normal)
        if row.get("foto_barang"):
            foto_path = os.path.join(
                current_app.root_path,
                "static",
                "uploads",
                row["foto_barang"]
            )
            if os.path.exists(foto_path):
                foto_cell = RLImage(foto_path, width=50, height=50)

        table_data.append([
            Paragraph(str(row.get("kode_klaim", "-")), normal),
            Paragraph(str(row.get("kode_laporan_kehilangan", "-")), normal),
            Paragraph(str(row.get("nama_barang", "-")), normal),
            Paragraph(str(row.get("nama_pelapor", "-")), normal),
            Paragraph(str(row.get("status", "-")), normal),
            Paragraph(str(row.get("deskripsi_khusus", "-")), normal),
            Paragraph(str(row.get("tanggal", "-")), normal),
            foto_cell
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[60, 60, 80, 80, 60, 110, 60, 50]
    )

    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'), 
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (5,1), (5,-1), 'LEFT'),  
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"laporan_klaim_{bulan_label.replace(' ', '_').lower()}.pdf",
        mimetype="application/pdf"
    )

@admin_bp.route('/klaim/export/excel')
def export_klaim_excel():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    bulan = request.args.get('bulan')
    if not bulan:
        return "Bulan tidak valid", 400

    bulan_label = format_bulan_indonesia(bulan)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            kode_laporan AS kode_klaim,
            kode_laporan_kehilangan AS kode_kehilangan,
            nama_barang,
            nama_pelapor,
            status,
            deskripsi_khusus,
            DATE_FORMAT(STR_TO_DATE(tanggal_lapor, '%Y-%m-%d'), '%d-%m-%Y') AS tanggal
        FROM klaim_barang
        WHERE DATE_FORMAT(STR_TO_DATE(tanggal_lapor, '%Y-%m-%d'), '%Y-%m') = %s
        ORDER BY STR_TO_DATE(tanggal_lapor, '%Y-%m-%d') ASC
    """, (bulan,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Klaim"

    ws.append([
        "Kode Klaim",
        "Kode Kehilangan",
        "Nama Barang",
        "Nama Pengklaim",
        "Status",
        "Deskripsi",
        "Tanggal"
    ])

    for d in data:
        ws.append([
            d['kode_klaim'],
            d['kode_kehilangan'],
            d['nama_barang'],
            d['nama_pelapor'],
            d['status'],
            d.get('deskripsi_khusus', '-'),
            d['tanggal']
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"laporan_klaim_{bulan_label.replace(' ', '_').lower()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@admin_bp.route('/beranda/hapus', methods=['POST'])
def hapus_laporan():
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    kode = data.get('kode')
    jenis = data.get('jenis')

    if not kode or not jenis:
        return jsonify({"success": False, "message": "Data tidak lengkap"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if jenis == "kehilangan":
            cursor.execute("DELETE FROM kehilangan WHERE kode_kehilangan=%s", (kode,))
        elif jenis == "penemuan":
            cursor.execute("SELECT gambar_barang FROM penemuan WHERE kode_barang=%s", (kode,))
            result = cursor.fetchone()
            if result and result[0]:
                foto_path = os.path.join('static_admin', 'upload', result[0])
                if os.path.exists(foto_path):
                    os.remove(foto_path)
            cursor.execute("DELETE FROM penemuan WHERE kode_barang=%s", (kode,))
        elif jenis == "klaim":
            cursor.execute("DELETE FROM klaim_barang WHERE kode_laporan=%s", (kode,))
        else:
            return jsonify({"success": False, "message": "Jenis laporan tidak dikenal"}), 400

        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "Kode tidak ditemukan"}), 404

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/kehilangan/daftar')
def daftar_kehilangan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT * FROM kehilangan
        WHERE is_arsip = 0
        ORDER BY tanggal_submit DESC, waktu_submit DESC
    """)
    kehilangan_list = cursor.fetchall()
    conn.close()

    return render_template(
        'daftar_kehilangan.html',
        kehilangan_list=kehilangan_list,
        role=session.get('role')
    )

@admin_bp.route('/kehilangan/tambah', methods=['GET', 'POST'])
def tambah_kehilangan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)

    if request.method == 'GET':
        kode_baru = generate_kode_kehilangan(cursor)
        cursor.close()
        conn.close()

        return render_template(
            'tambah_kehilangan.html',
            role=session.get('role'),
            kode_baru=kode_baru
        )

    kode_kehilangan = request.form.get('kode_kehilangan')
    nama_pelapor = request.form.get('nama_pelapor', '').strip()
    email = request.form.get('email', '').strip()
    no_telp = request.form.get('no_telp', '').strip()
    asal_negara = request.form.get('asal_negara', '').strip()
    kota = request.form.get('kota', '').strip()
    nama_barang = request.form.get('nama_barang', '').strip()
    kategori = request.form.get('kategori', '').strip()
    jenis_laporan = request.form.get('jenis_laporan', 'Kehilangan').strip()
    deskripsi = request.form.get('deskripsi', '').strip()
    catatan_admin = request.form.get('catatan_admin', '').strip()
    lokasi = request.form.get('lokasi', '').strip()
    tanggal_kehilangan = request.form.get('tanggal_kehilangan')
    status = request.form.get('status', 'Verifikasi').strip()

    required = [
        nama_pelapor, email, no_telp,
        nama_barang, kategori, deskripsi,
        lokasi, tanggal_kehilangan
    ]

    if any(not v for v in required):
        cursor.close()
        conn.close()
        return "Error: Semua field wajib diisi.", 400

    foto = request.files.get('foto')
    foto_filename = None
    MAX_MB = 5

    if foto and foto.filename:
        foto.seek(0, os.SEEK_END)
        file_size = foto.tell()
        foto.seek(0)

        upload_folder = os.path.join('static', 'uploads')
        os.makedirs(upload_folder, exist_ok=True)

        filename = secure_filename(foto.filename)
        foto_filename = datetime.now().strftime("%Y%m%d%H%M%S_") + filename
        save_path = os.path.join(upload_folder, foto_filename)

        if file_size > MAX_MB * 1024 * 1024:
            compress_image(foto, save_path)
        else:
            foto.save(save_path)

    now = datetime.now()
    tanggal_submit = now.strftime("%Y-%m-%d")
    waktu_submit = now.strftime("%H:%M")
    update_terakhir = now.strftime("%Y-%m-%d %H:%M")

    try:
        cursor.execute("""
            INSERT INTO kehilangan (
                kode_kehilangan,
                nama_pelapor,
                email,
                no_telp,
                asal_negara,
                kota,
                nama_barang,
                kategori,
                jenis_laporan,
                deskripsi,
                lokasi,
                tanggal_kehilangan,
                tanggal_submit,
                waktu_submit,
                update_terakhir,
                catatan,
                status,
                foto
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            kode_kehilangan,
            nama_pelapor,
            email,
            no_telp,
            asal_negara,
            kota,
            nama_barang,
            kategori,
            jenis_laporan,
            deskripsi,
            lokasi,
            tanggal_kehilangan,
            tanggal_submit,
            waktu_submit,
            update_terakhir,
            catatan_admin,
            status,
            foto_filename
        ))

        conn.commit()

    except Exception as ex:
        conn.rollback()
        cursor.close()
        conn.close()
        print("❌ Error INSERT kehilangan:", ex)
        return f"Terjadi kesalahan saat menyimpan: {ex}", 500

    cursor.close()
    conn.close()

    return redirect(url_for('admin_bp.daftar_kehilangan'))

@admin_bp.route('/kehilangan/detail')
def detail_kehilangan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    kode_kehilangan = request.args.get('kode')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT * FROM kehilangan WHERE kode_kehilangan = %s AND is_arsip = 0",
        (kode_kehilangan,)
    )
    laporan = cursor.fetchone()

    if not laporan:
        cursor.close()
        conn.close()
        return "Data tidak ditemukan", 404

    tanggal = laporan['tanggal_kehilangan']
    h_min_1 = tanggal - timedelta(days=1)
    h_plus_1 = tanggal + timedelta(days=1)

    cursor.execute("""
        SELECT 
            id,
            kode_barang,
            nama_barang,
            gambar_barang,
            tanggal_lapor
        FROM penemuan
        WHERE tanggal_lapor BETWEEN %s AND %s
        AND status_barang = 'Tersedia'
        ORDER BY tanggal_lapor ASC
    """, (h_min_1, h_plus_1))
    foto_penemuan = cursor.fetchall()

    cursor.execute("""
        SELECT kode_penemuan
        FROM rekomendasi_penemuan
        WHERE kode_kehilangan = %s
    """, (kode_kehilangan,))
    rekomendasi_ada = [row['kode_penemuan'] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    return render_template(
        'detail_kehilangan.html',
        laporan=laporan,
        foto_penemuan=foto_penemuan,
        rekomendasi_ada=rekomendasi_ada
    )

@admin_bp.route('/rekomendasi/simpan', methods=['POST'])
def simpan_rekomendasi():
    data = request.get_json()

    kode_kehilangan = data.get('kode_kehilangan')
    kode_penemuan_list = data.get('kode_penemuan_list')

    if not kode_kehilangan or not kode_penemuan_list:
        return jsonify({
            "success": False,
            "message": "Data tidak lengkap"
        }), 400

    admin = session.get('admin_nama', 'admin')

    conn = get_db_connection()
    cursor = conn.cursor()

    for kode_penemuan in kode_penemuan_list:
        cursor.execute("""
            INSERT INTO rekomendasi_penemuan
            (kode_kehilangan, kode_penemuan, dipilih_oleh)
            VALUES (%s, %s, %s)
        """, (kode_kehilangan, kode_penemuan, admin))

    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Rekomendasi berhasil disimpan"
    })
    
@admin_bp.route('/rekomendasi/list')
def list_rekomendasi():
    kode_kehilangan = request.args.get('kode_kehilangan')
    if not kode_kehilangan:
        return jsonify({"success": False, "rekomendasi": []})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT kode_penemuan
        FROM rekomendasi_penemuan
        WHERE kode_kehilangan = %s
    """, (kode_kehilangan,))
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    rekomendasi_list = [row['kode_penemuan'] for row in data]
    return jsonify({"success": True, "rekomendasi": rekomendasi_list})

def kirim_email_verifikasi(email_user, nama_pelapor, kode_kehilangan):
    try:
        print("🔥 STATUS VERIFIKASI TERDETEKSI")
        print("📧 EMAIL USER:", email_user)

        msg = Message(
            subject="Laporan Kehilangan Anda Telah Diverifikasi",
            recipients=[email_user]
        )

        msg.body = f"""
Halo {nama_pelapor},

Laporan kehilangan Anda dengan kode:
{kode_kehilangan}

Telah berhasil DIVERIFIKASI oleh petugas Lost & Found
Bandara Internasional Juanda.

Silakan pantau status laporan Anda secara berkala melalui website.

Terima kasih,
Admin Lost & Found Juanda
"""
        mail.send(msg)
        print("✅ EMAIL VERIFIKASI TERKIRIM KE:", email_user)

    except Exception as e:
        print("❌ GAGAL KIRIM EMAIL:", e)

@admin_bp.route('/api/kehilangan/update_status', methods=['POST'])
def update_status_kehilangan():
    data = request.get_json()

    kode = data.get('kode')
    status_baru = data.get('status')
    catatan = data.get('catatan', '')  

    if not kode or not status_baru:
        return jsonify({'success': False, 'message': 'Data tidak lengkap!'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT status, email, nama_pelapor
        FROM kehilangan
        WHERE kode_kehilangan = %s
    """, (kode,))
    lama = cursor.fetchone()

    if not lama:
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'message': 'Data tidak ditemukan'}), 404

    status_lama = lama['status']
    email_user = lama['email']
    nama_pelapor = lama['nama_pelapor']

    waktu_update = datetime.now().strftime("%Y-%m-%d %H:%M")

    cursor.execute("""
        UPDATE kehilangan
        SET status = %s,
            catatan = %s,
            update_terakhir = %s
        WHERE kode_kehilangan = %s
    """, (status_baru, catatan, waktu_update, kode))

    cursor.execute("""
        SELECT status, catatan
        FROM riwayat_status
        WHERE kode_kehilangan = %s
        ORDER BY waktu_update DESC
        LIMIT 1
    """, (kode,))
    last = cursor.fetchone()

    if not last or last['status'] != status_baru or last['catatan'] != catatan:
        cursor.execute("""
            INSERT INTO riwayat_status
            (kode_kehilangan, status, catatan, waktu_update)
            VALUES (%s, %s, %s, %s)
        """, (kode, status_baru, catatan, waktu_update))

    conn.commit()

    if status_baru == "Selesai":
        pindahkan_ke_arsip(kode, "kehilangan")

    cursor.close()
    conn.close()

    if status_baru == "Verifikasi" and status_lama != "Verifikasi":
        kirim_email_verifikasi(email_user, nama_pelapor, kode)

    return jsonify({
        'success': True,
        'message': 'Status dan catatan berhasil diperbarui!',
        'update_terakhir': waktu_update
    })

@admin_bp.route('/api/kehilangan/delete', methods=['POST'])
def api_hapus_kehilangan():
    data = request.get_json()
    kode = data.get("kode")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("DELETE FROM kehilangan WHERE kode_kehilangan = %s", (kode,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        print("Error delete:", e)
        return jsonify({"success": False})
    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/kehilangan/edit', methods=['GET', 'POST'])
def edit_kehilangan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    kode_kehilangan = request.args.get('kode')
    from_page = request.args.get('from', '')   

    if not kode_kehilangan:
        return "Kode kehilangan tidak ditemukan", 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'GET':
        cursor.execute("SELECT * FROM kehilangan WHERE kode_kehilangan=%s", (kode_kehilangan,))
        laporan = cursor.fetchone()

        lokasi_full = laporan.get('lokasi', '')

        if " - " in lokasi_full:
            terminal, tempat = lokasi_full.split(" - ", 1)
        else:
            terminal, tempat = lokasi_full, ""

        laporan['terminal'] = terminal
        laporan['tempat'] = tempat

        return render_template(
            'edit_kehilangan.html',
            laporan=laporan,
            from_page=from_page,         
            role=session.get('role')
        )

    elif request.method == 'POST':
        try:
            from_page = (
                request.args.get('from') or 
                request.form.get('from') or 
                ""
            )

            nama_pelapor = request.form['nama_pelapor']
            no_telp = request.form['no_telp']
            email = request.form['email']
            asal_negara = request.form['asal_negara']
            kota = request.form['kota']

            nama_barang = request.form['nama_barang']
            kategori = request.form['kategori']
            terminal = request.form['terminal']
            tempat = request.form['tempat']
            lokasi = f"{terminal} - {tempat}"
            deskripsi = request.form['deskripsi']
            catatan = request.form['catatan']
            status = request.form['status']

            update_terakhir = datetime.now().strftime("%Y-%m-%d %H:%M")

            cursor.execute("""
                UPDATE kehilangan
                SET nama_pelapor=%s,
                    no_telp=%s,
                    email=%s,
                    asal_negara=%s,
                    kota=%s,
                    nama_barang=%s,
                    kategori=%s,
                    lokasi=%s,
                    deskripsi=%s,
                    catatan=%s,
                    status=%s,
                    update_terakhir=%s
                WHERE kode_kehilangan=%s
            """, (
                nama_pelapor, no_telp, email, asal_negara, kota,
                nama_barang, kategori, lokasi, deskripsi, catatan,
                status, update_terakhir, kode_kehilangan
            ))

            conn.commit()
            cursor.close()
            conn.close()
            
            if from_page == "beranda":
                return redirect(url_for('admin_bp.beranda_admin'))

            elif from_page == "detail":
                return redirect(url_for('admin_bp.detail_kehilangan', kode=kode_kehilangan))

            else:
                return redirect(url_for('admin_bp.daftar_kehilangan'))

        except Exception as e:
            conn.rollback()
            cursor.close()
            conn.close()
            print("❌ Error saat update kehilangan:", e)
            return f"Terjadi kesalahan: {e}", 500

@admin_bp.route('/kehilangan/hapus/<int:id>', methods=['GET'])
def hapus_kehilangan(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("DELETE FROM kehilangan WHERE id = %s", (id,))
        conn.commit()
        print(f"✅ Data kehilangan dengan ID {id} berhasil dihapus.")
    except Exception as e:
        print("❌ Error saat menghapus data:", e)
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin_bp.daftar_kehilangan'))


@admin_bp.route('/penemuan/daftar')
def daftar_penemuan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT * FROM penemuan
        WHERE is_arsip = 0
        ORDER BY tanggal_lapor DESC, update_terakhir DESC
    """)
    penemuan_list = cursor.fetchall()
    conn.close()

    return render_template(
        'daftar_penemuan.html',
        penemuan_list=penemuan_list,
        role=session.get('role')
    )

@admin_bp.route('/penemuan/tambah', methods=['GET', 'POST'])
def tambah_penemuan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)

    if request.method == 'GET':
        kode_baru = generate_kode_penemuan(cursor)
        cursor.close()
        conn.close()
        return render_template('tambah_penemuan.html', kode_baru=kode_baru)

    nama_pelapor   = request.form.get('nama_pelapor', '').strip()
    no_telp        = request.form.get('no_telp', '').strip()
    email          = request.form.get('email', '').strip()
    nama_barang    = request.form.get('nama_barang', '').strip()
    kategori       = request.form.get('kategori', '').strip()
    lokasi         = request.form.get('lokasi', '').strip()
    deskripsi      = request.form.get('deskripsi', '').strip()
    tanggal_lapor  = request.form.get('tanggal_lapor', '').strip()

    kode_barang    = request.form.get('kode_barang')
    jenis_laporan  = request.form.get('jenis_laporan', 'Penemuan')
    status         = request.form.get('status', 'Verifikasi')
    status_barang  = request.form.get('status_barang', 'Tersedia')
    jenis_barang   = request.form.get('jenis_barang', 'Publik')

    if not all([kode_barang, nama_pelapor, nama_barang, kategori, lokasi, tanggal_lapor]):
        cursor.close()
        conn.close()
        return "Data wajib tidak boleh kosong!", 400

    foto = request.files.get('foto')
    foto_filename = None
    MAX_MB = 5

    if foto and foto.filename:
        foto.seek(0, os.SEEK_END)
        file_size = foto.tell()
        foto.seek(0)

        upload_folder = os.path.join('static', 'uploads')
        os.makedirs(upload_folder, exist_ok=True)

        filename = secure_filename(foto.filename)
        foto_filename = f"{int(time.time())}_{filename}"
        save_path = os.path.join(upload_folder, foto_filename)

        if file_size > MAX_MB * 1024 * 1024:
            compress_image(foto, save_path)
        else:
            foto.save(save_path)

    try:
        cursor.execute("""
            INSERT INTO penemuan (
                kode_barang, nama_pelapor, no_telp, email, nama_barang,
                kategori, jenis_laporan, lokasi, deskripsi, tanggal_lapor,
                update_terakhir, status, status_barang, gambar_barang, jenis_barang
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s,%s,%s,%s)
        """, (
            kode_barang, nama_pelapor, no_telp, email, nama_barang,
            kategori, jenis_laporan, lokasi, deskripsi, tanggal_lapor,
            status, status_barang, foto_filename, jenis_barang
        ))

        conn.commit()
        flash("Penemuan berhasil ditambahkan!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Terjadi kesalahan: {e}", "danger")
        return redirect(request.url)

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('admin_bp.daftar_penemuan'))

@admin_bp.route('/penemuan/detail')
def detail_penemuan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    kode = request.args.get('kode')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT * FROM penemuan WHERE kode_barang=%s AND is_arsip = 0",
        (kode,)
    )
    laporan = cursor.fetchone()
    conn.close()

    if not laporan:
        return "Data tidak ditemukan", 404

    return render_template('detail_penemuan.html', laporan=laporan, role=session.get('role'))

@admin_bp.route('/api/penemuan/update_status', methods=['POST'])
def update_status_penemuan():
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()

    kode = data.get('kode')
    status_baru = data.get('status')            
    status_barang = data.get('status_barang')  

    if not kode:
        return jsonify({
            "success": False,
            "message": "Kode tidak ada"
        }), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(buffered=True)

        update_terakhir = datetime.now().strftime("%Y-%m-%d %H:%M")

        query = """
            UPDATE penemuan
            SET update_terakhir = %s
        """
        params = [update_terakhir]

        if status_baru is not None:
            query += ", status = %s"
            params.append(status_baru)

        if status_barang is not None:
            query += ", status_barang = %s"
            params.append(status_barang)

        query += " WHERE kode_barang = %s"
        params.append(kode)

        cursor.execute(query, tuple(params))
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({
                "success": False,
                "message": "Kode tidak ditemukan"
            }), 404

        if status_baru == "Selesai":
            pindahkan_ke_arsip(kode, "penemuan")

        return jsonify({
            "success": True,
            "update_terakhir": update_terakhir
        })

    except Exception as e:
        print("❌ Error update_status_penemuan:", e)
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/klaim/baru', methods=['GET', 'POST'])
def tambah_klaim_penemuan():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)

    MAX_MB = 5

    if request.method == "GET":
        kode_barang = request.args.get("kode_barang")
        from_page = request.args.get("from")

        if not kode_barang:
            return "Kode barang tidak ditemukan", 400

        cursor.execute(
            "SELECT * FROM penemuan WHERE kode_barang = %s AND is_arsip = 0",
            (kode_barang,)
        )
        laporan = cursor.fetchone()

        cursor.close()
        conn.close()

        return render_template(
            'tambah_klaim_penemuan.html',
            laporan=laporan,
            from_page=from_page
        )

    nama_pelapor = request.form.get("nama", "").strip()
    no_telp = request.form.get("telp", "").strip()
    email = request.form.get("email", "").strip()
    deskripsi_khusus = request.form.get("deskripsiKhusus", "").strip()
    kode_barang = request.form.get("kodeBarang", "").strip()
    kode_laporan_kehilangan = request.form.get("kodeLaporanKehilangan", "").strip() or None
    from_page = request.form.get("from")

    tanggal_lapor = datetime.now().strftime("%Y-%m-%d")
    waktu_lapor = datetime.now().strftime("%H:%M")

    cursor.execute("SELECT nama_barang FROM penemuan WHERE kode_barang=%s", (kode_barang,))
    barang = cursor.fetchone()
    nama_barang = barang["nama_barang"]

    kode_baru = generate_kode_klaim(cursor)

    def allowed_file(filename):
        return filename.lower().endswith(('.jpg', '.jpeg', '.png'))

    def simpan_file(file_obj):
        if not file_obj or not file_obj.filename:
            return None

        if not allowed_file(file_obj.filename):
            return None

        upload_folder = os.path.join("static", "uploads")
        os.makedirs(upload_folder, exist_ok=True)

        filename = secure_filename(file_obj.filename)
        final_name = f"{int(time.time())}_{filename}"
        save_path = os.path.join(upload_folder, final_name)

        file_obj.seek(0, os.SEEK_END)
        size = file_obj.tell()
        file_obj.seek(0)

        if size > MAX_MB * 1024 * 1024:
            compress_image(file_obj, save_path)
        else:
            file_obj.save(save_path)

        return final_name

    path_identitas = simpan_file(request.files.get("foto_identitas"))
    path_foto = simpan_file(request.files.get("foto_barang"))
    path_bukti = simpan_file(request.files.get("bukti_laporan"))

    if not all([nama_pelapor, no_telp, email, deskripsi_khusus, kode_barang, path_identitas, path_foto]):
        flash("Data belum lengkap atau file tidak valid.", "warning")
        cursor.close()
        conn.close()
        return redirect(request.url)

    try:
        cursor.execute("""
            INSERT INTO klaim_barang (
                kode_laporan, kode_barang, kode_laporan_kehilangan,
                nama_barang, nama_pelapor, no_telp, email,
                deskripsi_khusus, identitas_diri, bukti_laporan, foto_barang,
                tanggal_lapor, waktu_lapor
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            kode_baru, kode_barang, kode_laporan_kehilangan,
            nama_barang, nama_pelapor, no_telp, email,
            deskripsi_khusus, path_identitas, path_bukti, path_foto,
            tanggal_lapor, waktu_lapor
        ))

        conn.commit()
        flash("Klaim berhasil ditambahkan", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Gagal menyimpan klaim: {e}", "danger")

    finally:
        cursor.close()
        conn.close()

    if from_page == "beranda":
        return redirect(url_for("admin_bp.beranda_admin"))

    return redirect(url_for("admin_bp.daftar_klaim_penemuan"))

@admin_bp.route('/penemuan/edit', methods=['GET', 'POST'])
def edit_penemuan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    kode = request.args.get('kode')
    from_page = request.args.get('from', 'daftar_penemuan')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'GET':
        cursor.execute("SELECT * FROM penemuan WHERE kode_barang=%s", (kode,))
        laporan = cursor.fetchone()
        conn.close()

        if not laporan:
            return "Data tidak ditemukan", 404

        return render_template(
            'edit_penemuan.html',
            laporan=laporan,
            role=session.get('role'),
            from_page=from_page
        )

    nama_pelapor = request.form['nama_pelapor']
    no_telp = request.form['no_telp']
    email = request.form['email']
    nama_barang = request.form['nama_barang']
    kategori = request.form['kategori']
    lokasi = request.form['lokasi']
    deskripsi = request.form['deskripsi']
    status = request.form['status']                 
    status_barang = request.form['status_barang']
    jenis_barang = request.form['jenis_barang']

    update_terakhir = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute("""
        UPDATE penemuan
        SET nama_pelapor=%s,
            no_telp=%s,
            email=%s,
            nama_barang=%s,
            kategori=%s,
            lokasi=%s,
            deskripsi=%s,
            status=%s,
            status_barang=%s,
            jenis_barang=%s,
            update_terakhir=%s
        WHERE kode_barang=%s
    """, (
        nama_pelapor, no_telp, email,
        nama_barang, kategori, lokasi,
        deskripsi, status, status_barang, jenis_barang,
        update_terakhir, kode
    ))

    conn.commit()
    conn.close()

    if from_page == "beranda":
        return redirect(url_for('admin_bp.beranda_admin'))

    return redirect(url_for('admin_bp.daftar_penemuan'))

@admin_bp.route('/penemuan/hapus', methods=['POST'])
def hapus_penemuan():
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    kode = data.get('kode')

    if not kode:
        return jsonify({"success": False, "message": "Kode tidak ditemukan"}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT gambar_barang FROM penemuan WHERE kode_barang=%s", (kode,))
        result = cursor.fetchone()
        if result and result[0]:
            foto_path = os.path.join('static_admin', 'upload', result[0])
            if os.path.exists(foto_path):
                os.remove(foto_path)

        cursor.execute("DELETE FROM penemuan WHERE kode_barang=%s", (kode,))
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "Kode tidak ditemukan"}), 404

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        cursor.close()
        conn.close()
        
@admin_bp.route("/api/penemuan/verify", methods=["POST"])
def verify_barang():
    try:
        data = request.get_json()
        kode = data.get("kode")

        if not kode:
            return jsonify({"success": False, "message": "Kode kosong"}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        query = """
            UPDATE penemuan
            SET status = 'Verifikasi',
                update_terakhir = NOW()
            WHERE kode_barang = %s
        """
        cursor.execute(query, (kode,))
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "Kode tidak ditemukan"}), 404

        return jsonify({"success": True, "message": "Berhasil diverifikasi"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        try:
            cursor.close()
            conn.close()
        except:
            pass

@admin_bp.route('/penemuan/klaim')
def daftar_klaim_penemuan():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT 
            k.id,
            k.kode_laporan AS kode_klaim,
            k.kode_barang,
            k.nama_pelapor,
            p.nama_barang,
            k.status,
            k.tanggal_lapor
        FROM klaim_barang k
        LEFT JOIN penemuan p ON k.kode_barang = p.kode_barang
        WHERE k.status NOT IN ('Selesai', 'Ditolak')
        ORDER BY k.tanggal_lapor DESC
    """)

    data_klaim = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template(
        "admin/klaim_penemuan.html",
        data_klaim=data_klaim
    )

@admin_bp.route('/penemuan/klaim/update_status', methods=['POST'])
def update_status_klaim():
    if not session.get('admin_logged_in'):
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    data = request.get_json()
    kode = data.get("kode_laporan")
    status_baru = data.get("status")
    catatan = data.get("catatan_admin") or ""

    if not kode or not status_baru:
        return jsonify({"success": False, "message": "Data tidak lengkap"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT kode_barang, kode_laporan_kehilangan
            FROM klaim_barang
            WHERE kode_laporan = %s
        """, (kode,))
        klaim = cursor.fetchone()

        if not klaim:
            return jsonify({"success": False, "message": "Klaim tidak ditemukan"}), 404

        kode_barang = klaim["kode_barang"]
        kode_kehilangan = klaim["kode_laporan_kehilangan"]

        cursor.execute("""
            UPDATE klaim_barang
            SET status = %s,
                catatan_admin = %s,
                update_terakhir = NOW()
            WHERE kode_laporan = %s
        """, (status_baru, catatan, kode))

        if status_baru.lower() in ["pending", "ditolak"]:
            status_barang = "Tersedia"
        elif status_baru.lower() in ["verifikasi", "siap diambil"]:
            status_barang = "Diklaim"
        else:
            status_barang = None

        if status_barang:
            cursor.execute("""
                UPDATE penemuan
                SET status_barang = %s,
                    update_terakhir = NOW()
                WHERE kode_barang = %s
            """, (status_barang, kode_barang))

            if status_baru.lower() == "ditolak":
                cursor.execute("""
                    UPDATE kehilangan k
                    JOIN klaim_barang c ON k.kode_kehilangan = c.kode_laporan_kehilangan
                    SET k.status = 'Barang Tidak Ditemukan',
                        k.catatan = 'Barang klaim Anda ditolak oleh admin',
                        k.update_terakhir = NOW()
                    WHERE c.kode_barang = %s
                """, (kode_barang,))

                conn.commit()

                try:
                    cursor.execute("""
                        SELECT DISTINCT k.kode_kehilangan
                        FROM kehilangan k
                        JOIN klaim_barang c ON k.kode_kehilangan = c.kode_laporan_kehilangan
                        WHERE c.kode_barang = %s
                    """, (kode_barang,))
                    semua_kehilangan = cursor.fetchall()
                    for row in semua_kehilangan:
                        pindahkan_ke_arsip(row['kode_kehilangan'], "kehilangan")
                except Exception as e:
                    print("❌ Error arsip kehilangan:", e)

                try:
                    pindahkan_ke_arsip(kode, "klaim_barang")
                except Exception as e:
                    print("❌ Error arsip klaim:", e)

            elif kode_kehilangan:
                if status_baru.lower() == "selesai":
                    cursor.execute("""
                        UPDATE kehilangan
                        SET status = 'Selesai',
                            update_terakhir = NOW()
                        WHERE kode_kehilangan = %s
                    """, (kode_kehilangan,))

                    conn.commit()
                    try:
                        pindahkan_ke_arsip(kode_kehilangan, "kehilangan")
                    except Exception as e:
                        print("❌ Error arsip kehilangan:", e)

                else:
                    cursor.execute("""
                        UPDATE kehilangan
                        SET status = 'Dalam Proses',
                            update_terakhir = NOW()
                        WHERE kode_kehilangan = %s
                    """, (kode_kehilangan,))

                if status_baru == "Selesai":
                    cursor.execute("""
                        UPDATE penemuan
                        SET status = 'Selesai',
                            status_barang = 'Selesai',
                            update_terakhir = NOW()
                        WHERE kode_barang = %s
                    """, (kode_barang,))

                    conn.commit()
                    try:
                        pindahkan_ke_arsip(kode_barang, "penemuan")
                    except Exception as e:
                        print("❌ Error arsip penemuan:", e)

                elif status_baru == "Ditolak":
                    conn.commit()
                    try:
                        pindahkan_ke_arsip(kode, "klaim_barang")
                    except Exception as e:
                        print("❌ Error arsip klaim:", e)

                else:
                    conn.commit()

                return jsonify({"success": True})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@admin_bp.route('/penemuan/klaim/detail/<kode_klaim>')
def detail_klaim_penemuan(kode_klaim):
    return render_template("admin/detail_klaim_penemuan.html", kode_klaim=kode_klaim)


@admin_bp.route('/penemuan/klaim/api')
def detail_klaim_penemuan_api():
    kode = request.args.get("kode")
    if not kode:
        return jsonify({"success": False, "message": "Kode kosong"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            k.*,
            k.kode_laporan_kehilangan,
            k.foto_barang AS foto_barang_klaim,
            k.bukti_laporan,
            p.nama_barang,
            p.kategori,
            p.lokasi,
            p.gambar_barang AS foto_barang_penemuan
        FROM klaim_barang k
        LEFT JOIN penemuan p ON k.kode_barang = p.kode_barang
        WHERE k.kode_laporan = %s
        LIMIT 1
    """, (kode,))
    
    data = cursor.fetchone()
    cursor.close()
    conn.close()

    if not data:
        return jsonify({"success": False, "message": "Data tidak ditemukan"}), 404
    
    if data.get("update_terakhir"):
        data["update_terakhir"] = data["update_terakhir"].strftime("%Y-%m-%d %H:%M")

    return jsonify({"success": True, "data": data})

@admin_bp.route('/penemuan/klaim/update', methods=['POST'])
def update_klaim_penemuan():
    data = request.get_json()
    kode = data["kode_laporan"]
    status = data["status"]
    catatan = data["catatan_admin"]

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE klaim_barang
        SET status = %s,
            catatan_admin = %s,
            update_terakhir = NOW()
        WHERE kode_laporan = %s
    """, (status, catatan, kode))

    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({"success": True})

@admin_bp.route('/arsip')
def arsip():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    auto_arsip_laporan()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    cursor.execute("SELECT * FROM arsip ORDER BY tanggal_arsip DESC")
    data_arsip = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template(
        'arsip.html',
        role=session.get('role'),
        data_arsip=data_arsip
    )

def auto_arsip_laporan():
    db = get_db_connection()
    cur = db.cursor(dictionary=True)

    batas_tanggal = datetime.now() - relativedelta(months=3)
    tanggal_inactive = datetime.now().replace(second=0, microsecond=0)

    cur.execute("""
        SELECT kode_kehilangan
        FROM kehilangan
        WHERE is_arsip = 0
          AND (
              tanggal_kehilangan < %s
              OR status IN ('Selesai', 'Barang Tidak Ditemukan')
          )
    """, (batas_tanggal,))

    for row in cur.fetchall():
        pindahkan_ke_arsip(row['kode_kehilangan'], 'kehilangan', tanggal_inactive)

    cur.execute("""
        SELECT kode_barang
        FROM penemuan
        WHERE is_arsip = 0
          AND tanggal_lapor < %s
    """, (batas_tanggal,))

    for row in cur.fetchall():
        pindahkan_ke_arsip(row['kode_barang'], 'penemuan', tanggal_inactive)

    cur.close()
    db.close()

def pindahkan_ke_arsip(kode, jenis_tabel, tanggal_inactive=None):
    db = get_db_connection()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT 1 FROM arsip WHERE kode=%s", (kode,))
    if cur.fetchone():
        cur.close()
        db.close()
        return False

    if jenis_tabel == "kehilangan":
        cur.execute("SELECT * FROM kehilangan WHERE kode_kehilangan=%s", (kode,))
        kolom_tanggal = "tanggal_kehilangan"

    elif jenis_tabel == "penemuan":
        cur.execute("SELECT * FROM penemuan WHERE kode_barang=%s", (kode,))
        kolom_tanggal = "tanggal_lapor"

    elif jenis_tabel in ("klaim", "klaim_barang"):
        cur.execute("""
            SELECT k.*, p.nama_barang, p.kategori, p.lokasi, p.gambar_barang
            FROM klaim_barang k
            LEFT JOIN penemuan p ON k.kode_barang = p.kode_barang
            WHERE k.kode_laporan=%s
        """, (kode,))
        kolom_tanggal = "tanggal_lapor"

    else:
        cur.close()
        db.close()
        return False

    data = cur.fetchone()
    if not data:
        cur.close()
        db.close()
        return False

    tanggal_asli = data.get(kolom_tanggal)
    tanggal_arsip = datetime.now().replace(second=0, microsecond=0)

    if jenis_tabel in ("klaim", "klaim_barang"):
        tanggal_inactive = None
    else:
        tanggal_inactive = None if data.get("status") in (
            "Selesai", "Barang Tidak Ditemukan"
        ) else tanggal_inactive

    cur.execute("""
        INSERT INTO arsip (
            kode, nama_barang, jenis, kategori, lokasi,
            tanggal, foto, nama_pelapor, no_telp, email,
            status, tanggal_arsip, tanggal_inactive
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        kode,
        data.get("nama_barang") or "-",
        "klaim" if jenis_tabel in ("klaim", "klaim_barang") else jenis_tabel,
        data.get("kategori") or "-",
        data.get("lokasi") or data.get("lokasi_kehilangan") or "-",
        tanggal_asli,
        data.get("gambar_barang") or data.get("foto") or "-",
        data.get("nama_pelapor") or data.get("nama_pengklaim") or "-",
        data.get("no_telp") or "-",
        data.get("email") or "-",
        data.get("status"),
        tanggal_arsip,
        tanggal_inactive
    ))

    if jenis_tabel == "kehilangan":
        cur.execute(
            "UPDATE kehilangan SET is_arsip=1 WHERE kode_kehilangan=%s",
            (kode,)
        )
    elif jenis_tabel == "penemuan":
        cur.execute(
            "UPDATE penemuan SET is_arsip=1 WHERE kode_barang=%s",
            (kode,)
        )
    else:
        cur.execute(
            "UPDATE klaim_barang SET is_arsip=1 WHERE kode_laporan=%s",
            (kode,)
        )

    db.commit()
    cur.close()
    db.close()
    return True

@admin_bp.route('/arsip/detail')
def detail_arsip():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    kode = request.args.get('kode')
    if not kode:
        return redirect(url_for('admin_bp.arsip'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)

    cursor.execute(
        "SELECT * FROM arsip WHERE kode=%s",
        (kode,)
    )
    arsip = cursor.fetchone()

    cursor.close()
    conn.close()

    if not arsip:
        flash("Data arsip tidak ditemukan", "danger")
        return redirect(url_for('admin_bp.arsip'))

    return render_template(
        'detail_arsip.html',
        data=arsip,
        role=session.get('role')
    )

def generate_kode_kehilangan(cursor):
    cursor.execute("SELECT kode_kehilangan FROM kehilangan")
    main_codes = [row['kode_kehilangan'] for row in cursor.fetchall() if row['kode_kehilangan']]

    cursor.execute("SELECT kode FROM arsip WHERE jenis='kehilangan'")
    archive_codes = [row['kode'] for row in cursor.fetchall() if row['kode']]

    all_codes = main_codes + archive_codes
    if not all_codes:
        return "LF-L001"

    max_num = max(
        int(re.search(r"LF-L(\d+)", code).group(1))
        for code in all_codes if re.search(r"LF-L(\d+)", code)
    )
    return f"LF-L{max_num + 1:03d}"

def generate_kode_penemuan(cursor):
    cursor.execute("SELECT kode_barang FROM penemuan")
    main_codes = [row['kode_barang'] for row in cursor.fetchall() if row['kode_barang']]

    cursor.execute("SELECT kode FROM arsip WHERE jenis='penemuan'")
    archive_codes = [row['kode'] for row in cursor.fetchall() if row['kode']]

    all_codes = main_codes + archive_codes
    if not all_codes:
        return "LF-F001"

    max_num = max(
        int(re.search(r"LF-F(\d+)", code).group(1))
        for code in all_codes if re.search(r"LF-F(\d+)", code)
    )
    return f"LF-F{max_num + 1:03d}"

def generate_kode_klaim(cursor):
    cursor.execute("SELECT kode_laporan FROM klaim_barang")
    main_codes = [row['kode_laporan'] for row in cursor.fetchall() if row['kode_laporan']]

    cursor.execute("SELECT kode FROM arsip WHERE jenis IN ('klaim', 'klaim barang')")
    archive_codes = [row['kode'] for row in cursor.fetchall() if row['kode']]

    all_codes = main_codes + archive_codes

    numbers = [int(re.search(r"LF-C(\d+)", code).group(1))
               for code in all_codes if re.search(r"LF-C(\d+)", code)]

    if not numbers:
        return "LF-C001"  

    max_num = max(numbers)
    return f"LF-C{max_num + 1:03d}"

@admin_bp.route('/pengaturan')
def pengaturan():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    cursor.execute("SELECT * FROM admin WHERE id=%s", (session['admin_id'],))
    admin_data = cursor.fetchone()
    cursor.close()
    conn.close()

    return render_template(
        'pengaturan.html',
        admin=admin_data,
        role=session.get('role')
    )

@admin_bp.route('/pengaturan/update', methods=['POST'])
def update_profile():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE admin SET full_name=%s, email=%s, phone=%s
        WHERE id=%s
    """, (
        request.form['full_name'],
        request.form['email'],
        request.form['phone'],
        session['admin_id']
    ))

    conn.commit()
    cursor.close()
    conn.close()

    session['admin_email'] = request.form['email']
    return redirect(url_for('admin_bp.pengaturan', updated=1))

@admin_bp.route('/pengaturan/upload-photo', methods=['POST'])
def upload_photo():
    if not session.get('admin_logged_in'):
        return {"status": "error"}, 401

    file = request.files.get('photo')
    if not file or not file.mimetype.startswith('image/'):
        return {"status": "error", "message": "File tidak valid"}, 400

    upload_folder = os.path.join(current_app.root_path, 'static', 'uploads')
    os.makedirs(upload_folder, exist_ok=True)

    filename = f"{session['admin_id']}_{secure_filename(file.filename)}"
    file.save(os.path.join(upload_folder, filename))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE admin SET photo=%s WHERE id=%s", (filename, session['admin_id']))
    conn.commit()
    cursor.close()
    conn.close()

    return {"status": "success", "filename": filename}

@admin_bp.app_context_processor
def inject_admin_profile():
    admin_data = None

    if 'admin_id' in session:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, full_name, photo FROM admin WHERE id = %s", (session['admin_id'],))
        admin_data = cursor.fetchone()
        cursor.close()
        conn.close()

    return {"global_admin": admin_data}

def is_super_admin():
    return session.get('role') == 'super_admin'

@admin_bp.route('/kelola_admin')
def kelola_admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    if not is_super_admin():
        return "Anda tidak memiliki akses!", 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM admin ORDER BY id DESC")
    admins = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template(
        'kelola_admin.html',
        admins=admins,
        role=session.get('role')   
    )

@admin_bp.route('/kelola_admin/tambah', methods=['GET', 'POST'])
def tambah_admin():
    if not is_super_admin():
        return "Akses ditolak", 403

    if request.method == 'GET':
        return render_template('tambah_admin.html')

    hashed_pw = generate_password_hash(request.form['password'])

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO admin (full_name, email, phone, password, role)
        VALUES (%s,%s,%s,%s,%s)
    """, (
        request.form['full_name'],
        request.form['email'],
        request.form.get('phone'),
        hashed_pw,
        request.form['role']
    ))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for('admin_bp.kelola_admin'))

@admin_bp.route('/kelola_admin/edit/<int:id>', methods=['GET', 'POST'])
def edit_admin(id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_bp.login_admin'))

    is_super = is_super_admin()
    is_self  = session.get('admin_id') == id

    if not is_super and not is_self:
        return "Anda tidak memiliki akses!", 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)

    cursor.execute("SELECT * FROM admin WHERE id=%s", (id,))
    target_admin = cursor.fetchone()

    if not target_admin:
        cursor.close()
        conn.close()
        return "Admin tidak ditemukan", 404

    if request.method == 'GET':
        cursor.close()
        conn.close()
        return render_template(
            'edit_admin.html',
            admin=target_admin,
            role=session.get('role')
        )

    full_name = request.form['full_name']
    email     = request.form['email']
    phone     = request.form.get('phone')
    password  = request.form.get('password')

    requested_role = request.form.get('role') if is_super else target_admin['role']

    if is_super and target_admin['role'] == "super_admin" and requested_role != "super_admin":
        cursor.execute("SELECT COUNT(*) AS total FROM admin WHERE role='super_admin'")
        total_super = cursor.fetchone()['total']
        if total_super <= 1:
            cursor.close()
            conn.close()
            return render_template(
                'edit_admin.html',
                admin=target_admin,
                role=session.get('role'),
                error="Minimal harus ada satu Super Admin."
            )

    final_role = requested_role

    try:
        if password:
            hashed_pw = generate_password_hash(password)

            cursor.execute("""
                UPDATE admin
                SET full_name=%s, email=%s, phone=%s, role=%s, password=%s
                WHERE id=%s
            """, (full_name, email, phone, final_role, hashed_pw, id))
        else:
            cursor.execute("""
                UPDATE admin
                SET full_name=%s, email=%s, phone=%s, role=%s
                WHERE id=%s
            """, (full_name, email, phone, final_role, id))

        conn.commit()

    except Exception as e:
        conn.rollback()
        print("❌ Error edit admin:", e)
        cursor.close()
        conn.close()
        return "Terjadi kesalahan", 500

    finally:
        cursor.close()
        conn.close()

    return redirect(
        url_for('admin_bp.kelola_admin') if is_super else url_for('admin_bp.pengaturan')
    )

@admin_bp.route('/pengaturan/ganti-password', methods=['GET', 'POST'])
def ganti_password():
    if not session.get('admin_logged_in'):
        return jsonify({
            "status": "error",
            "message": "Silakan login terlebih dahulu"
        }), 401

    if request.method == 'POST':
        admin_id = session.get('admin_id')

        old_pw = request.form.get('old_password', '').strip()
        new_pw = request.form.get('new_password', '').strip()
        confirm_pw = request.form.get('confirm_password', '').strip()

        if not old_pw or not new_pw or not confirm_pw:
            return jsonify({
                "status": "error",
                "message": "Semua field wajib diisi"
            }), 400

        if new_pw != confirm_pw:
            return jsonify({
                "status": "error",
                "message": "Konfirmasi password tidak cocok"
            }), 400

        if len(new_pw) < 8:
            return jsonify({
                "status": "error",
                "message": "Password minimal 8 karakter"
            }), 400

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True, buffered=True)

        cursor.execute(
            "SELECT password FROM admin WHERE id=%s",
            (admin_id,)
        )
        admin = cursor.fetchone()

        if not admin:
            cursor.close()
            conn.close()
            return jsonify({
                "status": "error",
                "message": "Data admin tidak ditemukan"
            }), 400

        if not check_password_hash(admin['password'], old_pw):
            cursor.close()
            conn.close()
            return jsonify({
                "status": "error",
                "message": "Password lama salah"
            }), 400

        hashed_pw = generate_password_hash(new_pw)

        cursor.execute(
            "UPDATE admin SET password=%s WHERE id=%s",
            (hashed_pw, admin_id)
        )
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({
            "status": "success",
            "message": "Password berhasil diperbarui"
        }), 200

    return render_template(
        'ganti_password.html',
        role=session.get('role')
    )

@admin_bp.route('/kelola_admin/delete/<int:id>', methods=['POST'])
def delete_admin(id):
    if not is_super_admin():
        return {"success": False}, 403

    if id == session['admin_id']:
        return {"success": False, "message": "Tidak bisa hapus diri sendiri"}

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM admin WHERE id=%s", (id,))
    conn.commit()
    cursor.close()
    conn.close()

    return {"success": True}

@admin_bp.route('/logout')
def logout_admin():
    session.clear()
    return redirect(url_for('admin_bp.login_admin'))
